# -*- coding: utf-8 -*-
"""导入服务：解析人工上传的 Excel/CSV 快照，写入独立库并留痕。

合规声明：本模块只读取【人工上传到本机】的静态文件，不发起任何网络请求。

官方表格结构（2026-08-16 实测）：
- 视频_机会排行.xlsx：sheet「机会排行」，R1=元信息行，R2 表头 [排名|搜索词|需求指数|机会指数]，510 条
- 视频_热搜排行.xlsx：sheet「热搜榜」+「上升榜」，R1=元信息行，R2 表头 [排名|搜索词|作品数]，各 100 条
"""
import csv
import json
import os
import re
import time

from db import get_conn, init_db, row_to_dict, save_upload_file, sha256_file

# 支持的中文/英文列名映射（含官方「搜索词/作品数/展示/转化/收益」列）
COLUMN_MAP = {
    # 题材/关键词/搜索词/视频标题列
    "theme": ["搜索词", "题材", "主题", "素材题材", "题材名", "主题词", "关键词", "热搜关键词", "热搜词",
              "theme", "topic", "subject", "keyword", "name", "名称", "题材名称"],
    "title": ["视频标题", "作品标题", "标题", "title", "video_title"],
    # 需求侧指数（含官方展示次数/热搜热度列）
    "demand": ["需求指数", "展示次数", "展示", "需求", "需求分", "指数", "热度", "搜索热度", "热搜指数", "热搜热度",
               "热度指数", "需求热度", "热力值", "demand", "demand_index", "heat", "hot", "popularity", "views", "impressions"],
    # 机会维度指数（官方机会表/关键词分析转化率）
    "opportunity": ["机会指数", "转化率", "机会", "机会分", "机会值", "机会度",
                    "opportunity", "opportunity_index", "opp", "conversion"],
    # 作品数（官方热搜表）
    "works": ["作品数", "作品数量", "作品", "works", "work_count", "works_count", "count"],
    "sales": ["销售次数", "销量", "销售数量", "数量", "sales", "qty", "sales_qty", "成交数"],
    "revenue": ["收益金额", "收益", "销售收入", "销售额", "销售额(元)", "销售金额", "金额",
                "revenue", "amount", "成交额"],
    "rank": ["排名", "榜单位次", "位次", "名次", "排行", "rank", "ranking", "no"],
}

# 官方表格类型识别
TABLE_TYPE_HINTS = {
    "opportunity_rank": ["机会指数", "需求指数", "opportunity"],
    "hot_keyword": ["作品数", "热搜", "热度", "works"],
    "keyword_analysis": ["展示次数", "点击次数", "转化率", "点击率"],
    "sales_record": ["购买时间", "订单编号", "收益金额", "购买账号"],
}
TABLE_TYPE_LABELS = {
    "opportunity_rank": "视频机会排行表（人工下载）",
    "hot_keyword": "视频热搜排行表（人工下载）",
    "keyword_analysis": "视频关键词分析周报（人工下载）",
    "sales_record": "自有视频销售记录（人工导出）",
    "generic": "通用快照",
}


def detect_table_type(headers) -> str:
    """根据表头识别官方表格类型：机会榜/热搜榜/关键词分析/销售记录"""
    norm = [_norm_header(h) for h in headers]
    for ttype in ("opportunity_rank", "hot_keyword", "keyword_analysis", "sales_record"):
        for hint in TABLE_TYPE_HINTS[ttype]:
            if _norm_header(hint) in norm:
                return ttype
    return "generic"


def _norm_header(h: str) -> str:
    return re.sub(r"[\s（）()【】\[\]：:]+", "", (h or "").strip().lower())


def _match_column(headers, key):
    """按列名映射找到目标列下标；返回 -1 表示未找到"""
    norm_headers = [_norm_header(h) for h in headers]
    for cand in COLUMN_MAP[key]:
        c = _norm_header(cand)
        if c in norm_headers:
            return norm_headers.index(c)
    return -1


def _to_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("，", "").replace("%", "")
    if s in ("", "-", "--", "N/A", "null", "None"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def _to_int(v):
    return int(_to_float(v))


def _find_header_row(rows):
    """定位表头行：官方表 R1 是元信息行（榜单时间/下载时间/时间范围），R2 才是表头；
    通用模板第一行即表头。策略：跳过元信息行（时间范围/下载/导出/榜单时间），
    找首个含数据列名（搜索词/关键词/题材/需求/机会/作品/购买时间/订单编号/收益金额/展示次数/转化率/标题）的行。"""
    meta_markers = ("时间范围", "下载时间", "下载文件时间", "导出时间", "榜单时间", "生成时间")
    data_markers = ("搜索词", "关键词", "题材", "需求", "机会", "作品", "热度", "指数", "主题",
                    "购买时间", "订单编号", "收益金额", "展示次数", "点击次数", "转化率", "视频标题", "标题", "销量", "销售额")
    for i, r in enumerate(rows[:6]):
        if not r:
            continue
        joined = "".join(str(v) for v in r if v is not None)
        if any(m in joined for m in meta_markers):
            continue
        if any(m in joined for m in data_markers):
            return i
    return 0


def _parse_rows(rows, sheet_name: str = ""):
    """按表头行解析数据行 → (records, table_type)"""
    if len(rows) < 2:
        return [], "generic"
    hdr_idx = _find_header_row(rows)
    headers = [str(h) if h is not None else "" for h in rows[hdr_idx]]
    table_type = detect_table_type(headers)
    idx = {
        "theme": _match_column(headers, "theme"),
        "title": _match_column(headers, "title"),
        "demand": _match_column(headers, "demand"),
        "opportunity": _match_column(headers, "opportunity"),
        "works": _match_column(headers, "works"),
        "sales": _match_column(headers, "sales"),
        "revenue": _match_column(headers, "revenue"),
        "rank": _match_column(headers, "rank"),
    }
    out = []
    for r in rows[hdr_idx + 1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        raw_theme = r[idx["theme"]] if idx["theme"] >= 0 and idx["theme"] < len(r) else None
        theme = str(raw_theme).strip() if raw_theme is not None else ""
        # 销售记录等：搜索词为空（含 None/空串）时回退视频标题
        if not theme and idx["title"] >= 0 and idx["title"] < len(r) and r[idx["title"]] is not None:
            theme = str(r[idx["title"]]).strip()
        if not theme:
            continue
        out.append({
            "theme_raw": theme,
            "demand_index": _to_float(r[idx["demand"]]) if idx["demand"] >= 0 else 0.0,
            "opportunity_index": _to_float(r[idx["opportunity"]]) if idx["opportunity"] >= 0 else 0.0,
            "works_count": _to_float(r[idx["works"]]) if idx["works"] >= 0 else 0.0,
            "sales_qty": _to_float(r[idx["sales"]]) if idx["sales"] >= 0 else (1.0 if table_type == "sales_record" else 0.0),
            "revenue": _to_float(r[idx["revenue"]]) if idx["revenue"] >= 0 else 0.0,
            "rank_no": _to_int(r[idx["rank"]]) if idx["rank"] >= 0 else 0,
            "sheet_name": sheet_name,
        })
    return out, table_type


def parse_excel(path: str, return_meta: bool = False):
    """解析人工 Excel 快照：遍历全部 sheet（官方热搜表含 热搜榜/上升榜 两个 sheet）
    return_meta=True 时返回 (records, table_type)；records 每条带 sheet_name/works_count"""
    import openpyxl
    # 注意：官方表 dimension 记录不完整，read_only=True 会截断行（512→9）；用普通模式读取
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    all_records = []
    table_type = "generic"
    sheets = []
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            recs, t = _parse_rows(rows, sheet_name=sn)
            if recs:
                sheets.append(sn)
                all_records.extend(recs)
                if t != "generic" and table_type == "generic":
                    table_type = t
    finally:
        wb.close()
    return (all_records, table_type) if return_meta else all_records


def parse_csv(path: str, return_meta: bool = False):
    """解析人工 CSV 快照（自动探测编码 UTF-8/GBK）；return_meta=True 时返回 (records, table_type)"""
    data = None
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                data = list(csv.reader(f))
            break
        except Exception:
            continue
    if data is None:
        raise ValueError("CSV 编码无法识别（支持 UTF-8 / GBK）")
    if not data:
        return ([], "generic") if return_meta else []
    recs, table_type = _parse_rows(data)
    return (recs, table_type) if return_meta else recs


def import_snapshot(file, file_name: str, source_type: str, version_name: str,
                    uploaded_by: str = "", note: str = "") -> dict:
    """人工上传快照入库：保存文件 → 解析（多 sheet）→ 写 raw_records + snapshot_versions → 留痕"""
    init_db()
    ext = os.path.splitext(file_name)[1].lower()
    dest_name = f"{int(time.time())}_{file_name.replace(os.sep, '_')}"
    path = save_upload_file(file, dest_name)
    fhash = sha256_file(path)

    try:
        if source_type == "csv" or ext == ".csv":
            parsed = parse_csv(path, return_meta=True)
        else:
            parsed = parse_excel(path, return_meta=True)
        records, table_type = parsed
    except Exception as e:
        _log_upload(file_name, source_type, 0, 0, 0, str(e), fhash, uploaded_by)
        raise ValueError(f"文件解析失败：{e}")

    if not records:
        _log_upload(file_name, source_type, 0, 0, 0, "无有效数据行（请检查表头：排名/搜索词/需求指数/机会指数/作品数）",
                    fhash, uploaded_by)
        raise ValueError("未解析到有效数据行，请检查表头（官方表：排名/搜索词/需求指数/机会指数/作品数）")

    sheets = sorted({r.get("sheet_name") or "" for r in records})
    type_label = TABLE_TYPE_LABELS.get(table_type, table_type)
    note_full = (note + " ｜ " if note else "") + f"表型识别：{type_label}"
    if sheets:
        note_full += f"｜sheets：{'、'.join(sheets)}"

    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO snapshot_versions (name, source_type, file_name, file_hash, rows_count, status, uploaded_by, note) "
            "VALUES (?,?,?,?,?,?,?,?)",
            [version_name or file_name, source_type, file_name, fhash, len(records), "imported", uploaded_by, note_full])
        vid = cur.lastrowid
        for r in records:
            conn.execute(
                "INSERT INTO raw_records (version_id, theme_raw, demand_index, opportunity_index, works_count, sales_qty, revenue, rank_no, sheet_name) "
                "VALUES (?,?,?,?,?,?,?,?,?)",
                [vid, r["theme_raw"], r["demand_index"], r["opportunity_index"], r["works_count"],
                 r["sales_qty"], r["revenue"], r["rank_no"], r.get("sheet_name") or ""])
        conn.commit()
        _log_upload(file_name, source_type, len(records), len(records), 0, "", fhash, uploaded_by)
        return {"version_id": vid, "rows": len(records), "file_hash": fhash[:16], "table_type": table_type, "sheets": sheets}
    finally:
        conn.close()


def _log_upload(file_name, source_type, rows_total, rows_ok, rows_fail, errors, fhash, uploaded_by):
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO upload_logs (file_name, source_type, rows_total, rows_ok, rows_fail, errors, file_hash, uploaded_by) "
            "VALUES (?,?,?,?,?,?,?,?)",
            [file_name, source_type, rows_total, rows_ok, rows_fail, errors[:500], fhash, uploaded_by])
        conn.commit()
    finally:
        conn.close()


def add_announcement(title: str, content: str, publish_date: str = "",
                     source_hint: str = "", entered_by: str = "") -> int:
    """官方公告人工录入（人工粘贴，无任何自动获取）"""
    init_db()
    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO announcements (title, content, publish_date, source_hint, entered_by) VALUES (?,?,?,?,?)",
            [title, content, publish_date, source_hint, entered_by])
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()
